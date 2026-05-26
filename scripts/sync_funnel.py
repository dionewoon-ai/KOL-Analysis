#!/usr/bin/env python3
"""
Sync funnel baseline + uplift metrics from local PMF Dashboard xlsx into data.js.

WHERE THE XLSX PATH IS SET:
  data.js → KOL_DATA.funnelSource.pmfDashboardXlsx
  (path is relative to this scripts/ folder, i.e. KOL Analysis/)

HOW BASELINE / UPLIFT ARE COMPUTED:
  See data.js → KOL_DATA.campaignTimeline.methodology

Usage (from KOL Analysis/):
  python3 scripts/sync_funnel.py
"""

from __future__ import annotations

import json
import re
import subprocess
import sys
from collections import defaultdict
from datetime import date, datetime, timedelta
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Install openpyxl: pip3 install openpyxl", file=sys.stderr)
    sys.exit(1)

ROOT = Path(__file__).resolve().parent.parent          # KOL Analysis/
DATA_JS = ROOT / "data.js"
DEFAULT_PARTNER_MAP = {
    "Drivelah": ["Drivelah"],
    "HelloRide": ["HelloRide"],
    "Jolibox": ["Jolibox Game", "Jolibox Drama"],
    "Redbus": ["RedBus"],
    "Firsty": ["Firsty"],
}


def load_kol_config() -> dict:
    """Evaluate data.js in Node and return funnelSource + campaignTimeline."""
    script = """
    const fs = require('fs');
    let src = fs.readFileSync(process.argv[1], 'utf8');
    src = src.replace(/^const KOL_DATA/m, 'var KOL_DATA');
    eval(src);
    console.log(JSON.stringify({
      funnelSource: KOL_DATA.funnelSource,
      campaignTimeline: KOL_DATA.campaignTimeline,
      partners: KOL_DATA.partners,
    }));
    """
    result = subprocess.run(
        ["node", "-e", script, str(DATA_JS)],
        capture_output=True,
        text=True,
        check=True,
    )
    return json.loads(result.stdout)


def parse_date(value) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        # Excel serial
        return (datetime(1899, 12, 30) + timedelta(days=float(value))).date()
    s = str(value).strip()
    if not s:
        return None
    return datetime.strptime(s[:10], "%Y-%m-%d").date()


def monday_of(d: date) -> date:
    return d - timedelta(days=d.weekday())


def baseline_week_starts(anchor: date, n: int = 4) -> list[date]:
    """Four complete Mon-start weeks immediately before anchor week."""
    anchor_monday = monday_of(anchor)
    return [anchor_monday - timedelta(weeks=i) for i in range(n, 0, -1)]


def uplift_week_starts(boost_start: date, boost_end: date) -> list[date]:
    """Mon-start weeks overlapping boostStart → boostEnd."""
    if boost_start > boost_end:
        return []
    first_monday = monday_of(boost_start)
    if first_monday + timedelta(days=6) < boost_start:
        first_monday += timedelta(weeks=1)
    weeks = []
    cur = first_monday
    while cur <= boost_end:
        if cur + timedelta(days=6) >= boost_start:
            weeks.append(cur)
        cur += timedelta(weeks=1)
    return weeks


def read_weekly_metrics(xlsx_path: Path, sheets: dict, partner_map: dict) -> dict:
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    funnel_rows = wb[sheets["funnelWeekly"]].iter_rows(values_only=True)
    funnel_header = [str(h).strip() if h is not None else "" for h in next(funnel_rows)]
    fi = {h: i for i, h in enumerate(funnel_header)}

    trend_rows = wb[sheets["weeklyTrend"]].iter_rows(values_only=True)
    trend_header = [str(h).strip() if h is not None else "" for h in next(trend_rows)]
    ti = {h: i for i, h in enumerate(trend_header)}

    weekly: dict[date, dict[str, dict]] = defaultdict(lambda: defaultdict(lambda: {"orders": 0.0, "gmv": 0.0, "weu": 0.0}))

    for row in funnel_rows:
        if not row or len(row) <= max(fi["partner_name"], fi["week_start"], fi["total_orders"], fi["gmv_usd"]):
            continue
        partner = row[fi["partner_name"]]
        if not partner or partner == "All Partners":
            continue
        ws = parse_date(row[fi["week_start"]])
        if ws is None:
            continue
        weekly[ws][str(partner)]["orders"] += float(row[fi["total_orders"]] or 0)
        weekly[ws][str(partner)]["gmv"] += float(row[fi["gmv_usd"]] or 0)

    for row in trend_rows:
        if not row or len(row) <= max(ti["partner_name"], ti["week_start"], ti["weus"]):
            continue
        partner = row[ti["partner_name"]]
        if not partner or partner == "All Partners":
            continue
        ws = parse_date(row[ti["week_start"]])
        if ws is None:
            continue
        weekly[ws][str(partner)]["weu"] += float(row[ti["weus"]] or 0)

    wb.close()

    # Roll up to KOL partner keys
    rolled: dict[date, dict[str, dict]] = defaultdict(lambda: defaultdict(lambda: {"orders": 0.0, "gmv": 0.0, "weu": 0.0}))
    for ws, pdata in weekly.items():
        for kol_partner, src_names in partner_map.items():
            for sn in src_names:
                if sn in pdata:
                    rolled[ws][kol_partner]["orders"] += pdata[sn]["orders"]
                    rolled[ws][kol_partner]["gmv"] += pdata[sn]["gmv"]
                    rolled[ws][kol_partner]["weu"] += pdata[sn]["weu"]
    return rolled


def avg_weeks(weekly: dict, partner: str, week_starts: list[date]) -> dict | None:
    if not week_starts:
        return None
    totals = {"orders": 0.0, "gmvUSD": 0.0, "weu": 0.0}
    n = 0
    for ws in week_starts:
        if ws not in weekly or partner not in weekly[ws]:
            continue
        d = weekly[ws][partner]
        totals["orders"] += d["orders"]
        totals["gmvUSD"] += d["gmv"]
        totals["weu"] += d["weu"]
        n += 1
    if n == 0:
        return None
    return {
        "weu": round(totals["weu"] / n, 1),
        "orders": round(totals["orders"] / n, 1),
        "gmvUSD": round(totals["gmvUSD"] / n, 1),
    }


def anchor_date(partner_cfg: dict) -> date:
    anchor_type = partner_cfg.get("baselineAnchor", "boost")
    if anchor_type == "post":
        d = parse_date(partner_cfg.get("postDate"))
    else:
        d = parse_date(partner_cfg.get("boostStart"))
    if d is None:
        raise ValueError(f"Missing anchor date for {partner_cfg}")
    return d


def compute_funnel(config: dict) -> dict:
    src = config["funnelSource"]
    timeline = config["campaignTimeline"]["partners"]
    partner_map = src.get("partnerMap") or DEFAULT_PARTNER_MAP

    xlsx_path = (ROOT / src["pmfDashboardXlsx"]).resolve()
    if not xlsx_path.exists():
        raise FileNotFoundError(f"PMF xlsx not found: {xlsx_path}\nUpdate funnelSource.pmfDashboardXlsx in data.js")

    weekly = read_weekly_metrics(xlsx_path, src["sheets"], partner_map)

    baseline_out, campaign_out, meta_out = {}, {}, {}

    for partner, pcfg in timeline.items():
        post = parse_date(pcfg.get("postDate"))
        boost_start = parse_date(pcfg.get("boostStart"))
        boost_end = parse_date(pcfg.get("boostEnd"))
        anchor = anchor_date(pcfg)

        base_weeks = baseline_week_starts(anchor, config["campaignTimeline"]["methodology"].get("baselineWeeks", 4))
        uplift_weeks = uplift_week_starts(boost_start, boost_end) if boost_start and boost_end else []

        baseline_out[partner] = avg_weeks(weekly, partner, base_weeks)
        campaign_out[partner] = avg_weeks(weekly, partner, uplift_weeks)

        meta_out[partner] = {
            "postDate": post.isoformat() if post else None,
            "boostStart": boost_start.isoformat() if boost_start else None,
            "boostEnd": boost_end.isoformat() if boost_end else None,
            "baselineAnchor": pcfg.get("baselineAnchor", "boost"),
            "baselineAnchorDate": anchor.isoformat(),
            "baselineWeeks": [w.isoformat() for w in base_weeks],
            "upliftWeeks": [w.isoformat() for w in uplift_weeks],
            "pmfSource": xlsx_path.name,
        }

    return {
        "baseline": baseline_out,
        "campaign": campaign_out,
        "meta": meta_out,
        "lastSynced": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "pmfDashboardXlsx": src["pmfDashboardXlsx"],
    }


def js_value(obj, indent=0) -> str:
    sp = "  " * indent
    if obj is None:
        return "null"
    if isinstance(obj, bool):
        return "true" if obj else "false"
    if isinstance(obj, (int, float)):
        if isinstance(obj, float) and obj == int(obj):
            return str(int(obj))
        return str(obj)
    if isinstance(obj, str):
        return json.dumps(obj)
    if isinstance(obj, list):
        if not obj:
            return "[]"
        lines = ["["]
        for item in obj:
            lines.append(f"{sp}  {js_value(item, indent + 1)},")
        lines.append(f"{sp}]")
        return "\n".join(lines)
    if isinstance(obj, dict):
        lines = ["{"]
        for k, v in obj.items():
            lines.append(f"{sp}  {json.dumps(k)}: {js_value(v, indent + 1)},")
        lines.append(f"{sp}}}")
        return "\n".join(lines)
    return json.dumps(obj)


def patch_data_js(funnel: dict) -> None:
    text = DATA_JS.read_text(encoding="utf-8")
    block = f"  funnel: {js_value(funnel, indent=1)},\n}};"
    new_text, n = re.subn(
        r"  funnel: \{[\s\S]*?\n\};",
        block,
        text,
        count=1,
    )
    if n != 1:
        raise RuntimeError("Could not locate funnel block in data.js")
    DATA_JS.write_text(new_text, encoding="utf-8")


def main() -> None:
    config = load_kol_config()
    funnel = compute_funnel(config)
    patch_data_js(funnel)
    print(f"Updated funnel in {DATA_JS}")
    print(f"Source: {funnel['pmfDashboardXlsx']}")
    for partner, m in funnel["meta"].items():
        print(f"  {partner}: baseline {m['baselineWeeks'][0]}..{m['baselineWeeks'][-1]} | uplift {m['upliftWeeks'][0] if m['upliftWeeks'] else '—'}..{m['upliftWeeks'][-1] if m['upliftWeeks'] else '—'}")


if __name__ == "__main__":
    main()
